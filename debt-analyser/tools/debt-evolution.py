import os
from collections import Counter
from openpyxl import load_workbook


def compare_files(file1, file2):
    """Compare two files and return the count of added and removed strings from .xlsx files."""
    wb1 = load_workbook(filename=file1)
    wb2 = load_workbook(filename=file2)
    
    # Assuming data is in the first sheet and in the first column
    ws1 = wb1.active
    ws2 = wb2.active
    
    # Extract strings from the first column
    strings1 = [row[0] for row in ws1.iter_rows(min_row=1, values_only=True) if row[0] is not None]
    strings2 = [row[0] for row in ws2.iter_rows(min_row=1, values_only=True) if row[0] is not None]
    
    # Use Counter to count occurrences of each string
    count1 = Counter(strings1)
    count2 = Counter(strings2)
    
    # Find differences
    added = sum((count2 - count1).values())
    removed = sum((count1 - count2).values())
    
    return added, removed, count1, count2


def process_directory(root_dir):
    """Process each directory and its files."""
    rem_dict = dict()
    for root, dirs, files in os.walk(root_dir):
        if len(dirs) == 0:  # This means we are in a leaf directory
            sorted_files = sorted(files, key=lambda x: int(x.split('.')[0]))  # Sort files by their numerical name
            for i in range(len(sorted_files) - 1):
                file1 = os.path.join(root, sorted_files[i])
                file2 = os.path.join(root, sorted_files[i + 1])
                added, removed, count1, _ = compare_files(file1, file2)
                if i == 0:
                    rem_dict[root] = [sum(count1.values()), 0]
                rem_dict[root] = [rem_dict[root][0] + added, rem_dict[root][1] + removed] 
                # print(f"Between {sorted_files[i]} and {sorted_files[i + 1]}: {added} added, {removed} removed in {root}")
    
    removed_instances = 0
    for k, v in rem_dict:
        if v[1] > 0:
            removed_instances += 1
    
    print(f'Occurence of debt removal: ${removed_instances * 100 / len(rem_dict.keys())}')


root_dir = '../debt_data/contracts'
process_directory(root_dir)